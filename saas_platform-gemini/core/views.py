import openpyxl
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
import jdatetime
from django.utils.timezone import localtime
from .models import ProductLog, Plan
from .forms import AddProductForm
from .services.access_control import AccessControlService
from .services.ai import AIService
from .services.woo import WooService

@login_required
def dashboard(request):
    user_profile = getattr(request.user, 'profile', None)
    if not user_profile:
        messages.error(request, "حساب کاربری شما به هیچ فروشگاهی متصل نیست.")
        return render(request, 'dashboard.html', {'error': True})

    store = user_profile.store
    remaining_quota = AccessControlService.get_remaining_quota(store)
    
    recent_add_logs = ProductLog.objects.filter(store=store, action_type='add').order_by('-created_at')[:15]
    recent_update_logs = ProductLog.objects.filter(store=store, action_type='update').order_by('-created_at')[:15]
    
    context = {
        'store': store,
        'remaining_quota': remaining_quota,
        'recent_add_logs': recent_add_logs,
        'recent_update_logs': recent_update_logs,
    }
    return render(request, 'dashboard.html', context)


@login_required
def clear_store_cache_view(request):
    user_profile = getattr(request.user, 'profile', None)
    if user_profile:
        woo_service = WooService(user_profile.store)
        woo_service.clear_store_cache()
        messages.success(request, "حافظه کِش دسته‌بندی‌های سایت شما با موفقیت تخلیه شد و در لود بعدی بروزرسانی می‌شود.")
    return redirect(request.META.get('HTTP_REFERER', 'dashboard'))


@login_required
def add_product(request):
    user_profile = getattr(request.user, 'profile', None)
    if not user_profile:
        return redirect('dashboard')

    store = user_profile.store
    if not AccessControlService.can_add_product(store):
        messages.error(request, "سقف مجاز افزودن محصول در ماه جاری برای پلن شما به اتمام رسیده است.")
        return redirect('dashboard')
        
    woo_service = WooService(store)
    categories_data = woo_service.get_categories()

    if request.method == 'POST':
        form = AddProductForm(request.POST, request.FILES, categories=categories_data)
        if form.is_valid():
            product_name = form.cleaned_data['product_name']
            price = form.cleaned_data['price']
            sku = form.cleaned_data['sku']
            category_id = form.cleaned_data['category_id']
            images = request.FILES.getlist('images')
            
            try:
                ai_service = AIService(store)
                slug = ai_service.generate_slug(product_name)
                short_desc = ai_service.generate_short_desc(product_name)
                long_desc = ai_service.generate_long_desc(product_name)
                
                if not long_desc:
                    raise Exception("هوش مصنوعی نتوانست محتوای مناسبی تولید کند.")

                image_ids = []
                for img in images:
                    img_id = woo_service.upload_image_to_wp(img.read(), img.name)
                    if img_id:
                        image_ids.append(img_id)
                        
                woo_product = woo_service.create_product(
                    title=product_name, price=price, sku=sku,
                    category_id=int(category_id) if category_id else None,
                    slug=slug, description=long_desc, short_description=short_desc, image_ids=image_ids
                )
                
                if woo_product:
                    ProductLog.objects.create(
                        store=store, user=request.user, product_name=product_name, 
                        woo_product_id=woo_product.get('id'), status='success', action_type='add',
                        details=f"محصول جدید با قیمت {price} تومان و شناسه یکتای {sku or '---'} ثبت و منتشر شد."
                    )
                    messages.success(request, f"محصول «{product_name}» با موفقیت منتشر شد.")
                else:
                    raise Exception("خطا در ثبت نهایی محصول در ووکامرس.")
                    
            except Exception as e:
                ProductLog.objects.create(store=store, user=request.user, product_name=product_name, status='failed', action_type='add', error_message=str(e))
                messages.error(request, f"خطا: {str(e)}")
            return redirect('add_product')
    else:
        form = AddProductForm(categories=categories_data)
        
    return render(request, 'add_product.html', {'form': form, 'store': store})


@login_required
def product_manager(request):
    user_profile = getattr(request.user, 'profile', None)
    if not user_profile:
        return redirect('dashboard')

    store = user_profile.store
    if not AccessControlService.has_feature(store, 'price_manager'):
        messages.error(request, "پلن اشتراکی شما به این بخش دسترسی ندارد.")
        return redirect('dashboard')

    woo_service = WooService(store)
    categories_data = woo_service.get_categories()

    # تغییر قیمت گروهی
    if request.method == 'POST' and 'bulk_update' in request.POST:
        bulk_category = request.POST.get('bulk_category')
        action = request.POST.get('action') 
        percent = float(request.POST.get('percent', 0))
        
        if bulk_category and percent > 0:
            updated_count = woo_service.batch_update_prices(bulk_category, percent, action)
            if updated_count > 0:
                action_text = "افزایش درصدی" if action == "increase" else "کاهش درصدی"
                ProductLog.objects.create(
                    store=store, user=request.user, product_name=f"تغییر قیمت گروهی دسته {bulk_category}",
                    status='success', action_type='update',
                    details=f"تعداد {updated_count} محصول در دسته {bulk_category} تحت عملیات {action_text} با مقدار {percent}% قرار گرفتند."
                )
                messages.success(request, "قیمت‌ها با موفقیت تغییر کردند.")
            else:
                messages.warning(request, "محصولی پیدا نشد.")
        return redirect('product_manager')

    # ویرایش تکی محصول
    if request.method == 'POST' and 'single_update' in request.POST:
        product_id = request.POST.get('product_id')
        reg_price = request.POST.get('regular_price', '').strip()
        sale_price = request.POST.get('sale_price', '').strip()
        stock_status = request.POST.get('stock_status')
        
        payload = {
            "stock_status": stock_status
        }
        
        if reg_price:
            payload["regular_price"] = str(reg_price)
            
        if sale_price:
            payload["sale_price"] = str(sale_price)
        else:
            payload["sale_price"] = "" # حذف حراجی در صورت خالی بودن

        # 🔴 ترفند مهم برای رفع ارور ووکامرس هنگام ناموجود کردن محصول
        if stock_status == 'outofstock':
            payload['manage_stock'] = False
            payload['stock_quantity'] = 0
            
        result = woo_service.update_product(product_id, payload)
        if result:
            stock_text = "موجود" if stock_status == "instock" else ("پیش‌سفارش" if stock_status == "onbackorder" else "ناموجود")
            ProductLog.objects.create(
                store=store, user=request.user, product_name=result.get('name', 'محصول تکی'),
                woo_product_id=product_id, status='success', action_type='update',
                details=f"تغییرات تکی: قیمت اصلی: {reg_price or '---'} تومان | قیمت حراج: {sale_price or '---'} تومان | وضعیت موجودی: {stock_text}"
            )
            messages.success(request, f"محصول «{result.get('name')}» بروزرسانی شد.")
        else:
            messages.error(request, "خطا در بروزرسانی محصول.")
        return redirect('product_manager')

    search_q = request.GET.get('search', '')
    sku_q = request.GET.get('sku', '')
    cat_q = request.GET.get('category', '')
    stock_q = request.GET.get('stock_status', '')
    page = int(request.GET.get('page', 1))

    products, total_pages = woo_service.get_products_filtered(
        page=page, per_page=20, search=search_q, category=cat_q, stock_status=stock_q, sku=sku_q
    )

    context = {
        'store': store, 'products': products, 'categories': categories_data,
        'current_page': page, 'total_pages': total_pages, 'page_range': range(1, total_pages + 1),
        'search_q': search_q, 'sku_q': sku_q, 'cat_q': cat_q, 'stock_q': stock_q,
    }
    return render(request, 'product_manager.html', context)


@login_required
def export_logs_excel(request):
    user_profile = getattr(request.user, 'profile', None)
    if not user_profile:
        return redirect('dashboard')

    store = user_profile.store
    logs = ProductLog.objects.filter(store=store).select_related('user').order_by('-created_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "گزارش فعالیت‌ها"
    ws.sheet_view.rightToLeft = True

    headers = ["ردیف", "نام محصول", "آیدی ووکامرس", "کاربر ثبت‌کننده", "نوع عملیات", "وضعیت", "جزئیات تکمیلی تغییرات", "تاریخ شمسی", "ساعت"]
    ws.append(headers)

    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F8EF7")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for idx, log in enumerate(logs, 1):
        local_dt = localtime(log.created_at)
        jd = jdatetime.datetime.fromgregorian(datetime=local_dt)
        
        ws.append([
            idx, log.product_name, log.woo_product_id or "---", log.user.username if log.user else "سیستم",
            log.get_action_type_display(), "موفق" if log.status == 'success' else "ناموفق",
            log.details or log.error_message or "---", jd.strftime("%Y/%m/%d"), jd.strftime("%H:%M")
        ])

    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['G'].width = 50

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="SaaS_Report_{store.name}.xlsx"'
    wb.save(response)
    return response

def pricing_view(request):
    """
    ویوی نمایش پلن‌های اشتراکی سیستم
    """
    plans = Plan.objects.filter(is_active=True)
    return render(request, 'pricing.html', {'plans': plans})